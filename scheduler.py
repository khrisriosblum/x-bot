# scheduler.py
from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any

import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

from settings import settings
from post_generator import build_post   # alias a build_copy
from x_client import XClient

logger = logging.getLogger("xbot")


# ---------------------------
# Helpers horario
# ---------------------------
def _parse_slot_to_hm(slot: str) -> tuple[int, int]:
    """
    Acepta 'HH' o 'HH:MM' y devuelve (hora, minuto).
    Lanza ValueError si es inválido.
    """
    parts = slot.strip().split(":")
    if len(parts) == 1:
        h, m = int(parts[0]), 0
    elif len(parts) == 2:
        h, m = int(parts[0]), int(parts[1])
    else:
        raise ValueError(f"Formato de slot inválido '{slot}'. Usa 'HH' o 'HH:MM'.")
    if not (0 <= h <= 23 and 0 <= m <= 59):
        raise ValueError(f"Hora/minuto fuera de rango en slot '{slot}'")
    return h, m


# ---------------------------
# Estructura de candidato
# ---------------------------
@dataclass
class Candidate:
    row_idx: int
    platform: str
    url_col: str
    posted_col: str
    last_posted_col: str
    url: str
    title: str
    artist: Optional[str]
    lang: Optional[str]
    release_dt: Optional[datetime]
    last_posted_at: Optional[datetime]
    is_recent: bool


# ---------------------------
# BotScheduler
# ---------------------------
class BotScheduler:
    def __init__(self) -> None:
        self.scheduler = BackgroundScheduler(timezone=ZoneInfo(settings.TZ))
        self.tz = ZoneInfo(settings.TZ)
        self.jitter_seconds = int(getattr(settings, "SLOT_JITTER_MINUTES", 15)) * 60

        # Mapa columnas por plataforma
        self.platform_cols: Dict[str, Dict[str, str]] = {
            "YouTube":    {"url": "YouTubeURL",    "posted": "PostedYouTube",    "last": "LastPostedYouTubeAt"},
            "Beatport":   {"url": "BeatportURL",   "posted": "PostedBeatport",   "last": "LastPostedBeatportAt"},
            "AppleMusic": {"url": "AppleMusicURL", "posted": "PostedAppleMusic", "last": "LastPostedAppleMusicAt"},
            "Spotify":    {"url": "SpotifyURL",    "posted": "PostedSpotify",    "last": "LastPostedSpotifyAt"},
        }

    # ---- Arranque del scheduler
    def start(self) -> None:
        # Programa un job por cada slot
        for i, slot in enumerate(settings.DAILY_SLOTS):
            h, m = _parse_slot_to_hm(slot)
            trig = CronTrigger(hour=h, minute=m, timezone=self.tz)
            self.scheduler.add_job(
                self.run_once,
                trigger=trig,
                id=f"slot-{i}",
                coalesce=True,
                max_instances=1,
                misfire_grace_time=30 * 60,
                jitter=self.jitter_seconds,  # retraso aleatorio 0..jitter
            )
            logger.info({"event": "slot_scheduled", "slot": slot, "id": f"slot-{i}", "jitter_sec": self.jitter_seconds})
        self.scheduler.start()
        logger.info({"event": "scheduler_started", "tz": str(self.tz)})

    # ---- Lanza un post ahora (usado por el endpoint opcional)
    def post_job_with_jitter(self, force_no_jitter: bool = True) -> None:
        # Ejecutamos directo; el jitter ya se aplica sólo en cron.
        self.run_once()

    # ---- Job principal (una publicación)
    def run_once(self) -> None:
        now = datetime.now(self.tz)
        try:
            logger.info({"event": "job_start", "ts": now.isoformat()})
            df = self._load_tracks_df()
            if df is None or df.empty:
                logger.warning({"event": "no_data"})
                return

            candidates = self._explode_candidates(df, now)
            if not candidates:
                logger.warning({"event": "no_candidates"})
                return

            cand = self._pick_candidate(candidates, now)
            if not cand:
                logger.warning({"event": "no_candidate_after_filters"})
                return

            # Generar copy
            text = build_post(
                title=cand.title,
                artist=cand.artist,
                lang=cand.lang,
                release_dt=cand.release_dt,
                platform=cand.platform,
                url=cand.url,
            )

            # Publicar
            x = XClient()
            media_ids = x.prepare_thumbnail_if_enabled(cand.url, cand.platform)
            resp = x.post_text(text, media_ids=media_ids)

            logger.info({"event": "post_done", "platform": cand.platform, "url": cand.url, "dry_run": settings.DRY_RUN, "resp": resp})

            # Persistencia (si NO es dry-run)
            if not settings.DRY_RUN:
                self._mark_posted(df, cand, now)
                self._save_tracks_df(df)

        except Exception as e:
            logger.exception({"event": "job_exception", "error": str(e)})
        finally:
            logger.info({"event": "job_end", "ts": datetime.now(self.tz).isoformat()})

    # ---------------------------
    # Excel helpers
    # ---------------------------
    def _load_tracks_df(self) -> Optional[pd.DataFrame]:
        try:
            df = pd.read_excel(settings.EXCEL_PATH, sheet_name=getattr(settings, "EXCEL_SHEET", "Tracks"))
        except FileNotFoundError:
            logger.error({"event": "excel_not_found", "path": settings.EXCEL_PATH})
            return None
        except Exception as e:
            logger.error({"event": "excel_read_error", "error": str(e)})
            return None

        # Normaliza columnas básicas
        base_cols = ["Title", "Artist", "Language", "ReleaseDate"]
        for c in base_cols:
            if c not in df.columns:
                df[c] = None

        # Fechas
        if "ReleaseDate" in df.columns:
            df["ReleaseDate"] = pd.to_datetime(df["ReleaseDate"], errors="coerce")

        # Flags por plataforma
        for plat, cols in self.platform_cols.items():
            if cols["url"] not in df.columns:
                df[cols["url"]] = None
            if cols["posted"] not in df.columns:
                df[cols["posted"]] = False
            if cols["last"] not in df.columns:
                df[cols["last"]] = pd.NaT
            else:
                df[cols["last"]] = pd.to_datetime(df[cols["last"]], errors="coerce")

        return df

    def _save_tracks_df(self, df: pd.DataFrame) -> None:
        path = settings.EXCEL_PATH
        sheet = getattr(settings, "EXCEL_SHEET", "Tracks")

        try:
            wb = load_workbook(path)
            if sheet in wb.sheetnames:
                # Elimina la hoja para reemplazarla
                ws = wb[sheet]
                wb.remove(ws)
            # Escribe preservando otras hojas
            with pd.ExcelWriter(path, engine="openpyxl", mode="a") as writer:
                writer.book = wb
                writer.sheets = {ws.title: ws for ws in wb.worksheets}
                df.to_excel(writer, sheet_name=sheet, index=False)
                writer.save()
            logger.info({"event": "excel_saved", "path": path, "sheet": sheet})
        except FileNotFoundError:
            # Si no existía, crea el archivo
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
            logger.info({"event": "excel_created", "path": path, "sheet": sheet})
        except Exception as e:
            logger.error({"event": "excel_write_error", "error": str(e)})

    # ---------------------------
    # Selección de candidato
    # ---------------------------
    def _explode_candidates(self, df: pd.DataFrame, now: datetime) -> List[Candidate]:
        plats = settings.platforms  # lista desde PLATFORMS_ENABLED
        out: List[Candidate] = []
        recent_cut = now - timedelta(days=int(getattr(settings, "RECENT_DAYS_PRIORITY", 60)))

        for row_idx, row in df.iterrows():
            title = str(row.get("Title") or "").strip()
            artist = (str(row.get("Artist")) if pd.notna(row.get("Artist")) else None)
            lang = (str(row.get("Language")) if pd.notna(row.get("Language")) else None)
            rdt = row.get("ReleaseDate")
            rdt_dt: Optional[datetime] = pd.to_datetime(rdt, errors="coerce").to_pydatetime() if pd.notna(rdt) else None

            for plat in plats:
                if plat not in self.platform_cols:
                    continue
                cols = self.platform_cols[plat]
                url = row.get(cols["url"])
                if pd.isna(url) or not str(url).strip():
                    continue

                last_at = row.get(cols["last"])
                last_dt = pd.to_datetime(last_at, errors="coerce").to_pydatetime() if pd.notna(last_at) else None

                cand = Candidate(
                    row_idx=row_idx,
                    platform=plat,
                    url_col=cols["url"],
                    posted_col=cols["posted"],
                    last_posted_col=cols["last"],
                    url=str(url).strip(),
                    title=title,
                    artist=artist,
                    lang=lang,
                    release_dt=rdt_dt,
                    last_posted_at=last_dt,
                    is_recent=(rdt_dt is not None and rdt_dt >= recent_cut),
                )

                # Filtro de cooldown por URL (no repetir el MISMO URL en <= COOLDOWN días)
                if self._is_url_in_cooldown(df, cand.url, plat, now):
                    continue

                out.append(cand)

        return out

    def _is_url_in_cooldown(self, df: pd.DataFrame, url: str, platform: str, now: datetime) -> bool:
        cols = self.platform_cols[platform]
        cooldown_days = int(getattr(settings, "COOLDOWN_DAYS_PER_URL", 30))
        cutoff = now - timedelta(days=cooldown_days)

        # Busca en TODAS las filas donde el URL de esa plataforma coincida
        mask = (df[cols["url"]].astype(str).str.strip() == url.strip())
        if mask.any():
            last_col = cols["last"]
            # Si existe algún last_posted >= cutoff, bloquea
            recent = pd.to_datetime(df.loc[mask, last_col], errors="coerce")
            if recent.notna().any() and (recent >= cutoff).any():
                return True
        return False

    def _pick_candidate(self, candidates: List[Candidate], now: datetime) -> Optional[Candidate]:
        if not candidates:
            return None

        # Orden: recientes primero, luego por fecha de lanzamiento (desc), y como desempate, índice original
        def sort_key(c: Candidate):
            # Para None en release_dt, poner muy antiguo
            rd = c.release_dt or datetime(1970, 1, 1, tzinfo=self.tz)
            return (0 if c.is_recent else 1, -rd.timestamp(), c.row_idx)

        candidates.sort(key=sort_key)
        return candidates[0]

    # ---------------------------
    # Persistencia tras publicar
    # ---------------------------
    def _mark_posted(self, df: pd.DataFrame, cand: Candidate, when: datetime) -> None:
        df.loc[cand.row_idx, cand.posted_col] = True
        df.loc[cand.row_idx, cand.last_posted_col] = when


# Sugerencia: si tu main.py necesita una instancia global para endpoints, puedes crearla así:
# scheduler_instance = BotScheduler()
